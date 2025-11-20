from django.core.management.base import BaseCommand
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class Command(BaseCommand):
    help = 'Genera un archivo Excel de ejemplo para importar calificaciones'

    def handle(self, *args, **options):
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Calificaciones"

        # Definir estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Títulos de columnas
        headers = ["materia_codigo", "valor", "tipo", "peso", "comentario"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Datos de ejemplo
        datos_ejemplo = [
            ["MAT101", 4.5, "Parcial 1", 20, "Buen desempeño"],
            ["MAT101", 4.2, "Parcial 2", 20, "Satisfactorio"],
            ["MAT101", 4.8, "Examen Final", 30, "Excelente"],
            ["ING201", 3.9, "Quiz", 10, "Aceptable"],
            ["ING201", 4.1, "Proyecto", 30, "Muy bien"],
            ["FIS101", 3.5, "Laboratorio", 25, "Necesita mejorar"],
        ]

        # Insertar datos
        for row_num, row_data in enumerate(datos_ejemplo, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 25

        # Guardar archivo
        wb.save("Ejemplo_Importar_Calificaciones.xlsx")
        self.stdout.write(self.style.SUCCESS('✓ Archivo creado: Ejemplo_Importar_Calificaciones.xlsx'))

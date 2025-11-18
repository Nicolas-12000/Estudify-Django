from datetime import datetime
from io import BytesIO

import pandas as pd
from django.db.models import Avg
from django.http import HttpResponse
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)


class PDFReportGenerator:
    """
    Generador de reportes en PDF usando ReportLab.
    """

    @staticmethod
    def generate_grade_report(student, grades, course=None):
        """
        Generar boletín de calificaciones para un estudiante.

        Args:
            student: Usuario estudiante
            grades: QuerySet de calificaciones
            course: Curso específico (opcional)

        Returns:
            HttpResponse con el PDF generado
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#FF0000'),
            spaceAfter=30,
            alignment=1  # Centrado
        )

        title = Paragraph("BOLETÍN DE CALIFICACIONES", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        # Información del estudiante
        info_style = styles['Normal']
        student_info = [
            f"<b>Estudiante:</b> {student.get_full_name()}",
            f"<b>Código: </b> {student.username}",
            f"<b>Email: </b> {student.email}",
            f"<b>Fecha: </b> {datetime.now().strftime('%d/%m/%Y')}",
        ]

        for info in student_info:
            elements.append(Paragraph(info, info_style))

        elements.append(Spacer(1, 0.3 * inch))

        # Tabla de calificaciones
        data = [['Materia', 'Tipo', 'Calificación', 'Peso %', 'Estado']]

        for grade in grades:
            data.append([
                grade.subject.name,
                grade.get_grade_type_display(),
                str(grade.value),
                str(grade.weight),
                'Aprobado' if grade.is_passing else 'Reprobado'
            ])

        # Calcular promedios
        if grades:
            average = grades.aggregate(Avg('value'))['value__avg']
            data.append(['', '', '', 'PROMEDIO:', f"{average: .2f}"])

        table = Table(
            data,
            colWidths=[
                2.5 * inch,
                1.5 * inch,
                1 * inch,
                1 * inch,
                1 * inch,
            ],
        )
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF0000')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Cuerpo
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('GRID', (0, 0), (-1, -2), 1, colors.black),

            # Fila de promedio
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#333333')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))

        elements.append(table)

        # Construir PDF
        doc.build(elements)
        buffer.seek(0)

        # Crear respuesta HTTP
        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f'boletin_{student.username}_{datetime.now().strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = 'attachment' + '\x3B' + f' filename="{filename}"'

        return response

    @staticmethod
    def generate_attendance_report(student, attendances, course):
        """
        Generar reporte de asistencia para un estudiante.
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Título
        title = Paragraph("REPORTE DE ASISTENCIA", styles['Heading1'])
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        # Info
        info = [
            f"<b>Estudiante:</b> {student.get_full_name()}",
            f"<b>Curso: </b> {course.name}",
            f"<b>Fecha: </b> {datetime.now().strftime('%d/%m/%Y')}",
        ]

        for i in info:
            elements.append(Paragraph(i, styles['Normal']))

        elements.append(Spacer(1, 0.3 * inch))

        # Tabla de asistencias
        data = [['Fecha', 'Estado', 'Notas']]

        for att in attendances:
            data.append([
                att.date.strftime('%d/%m/%Y'),
                att.get_status_display(),
                att.notes or '-'
            ])

        # Estadísticas
        total = attendances.count()
        present = attendances.filter(status='PRESENT').count()
        absent = attendances.filter(status='ABSENT').count()
        late = attendances.filter(status='LATE').count()

        attendance_rate = ((present + late) / total * 100) if total > 0 else 0

        data.append(['', '', ''])
        data.append(['ESTADÍSTICAS', '', ''])
        data.append(['Total registros:', str(total), ''])
        data.append(['Asistencias:', str(present), ''])
        data.append(['Ausencias:', str(absent), ''])
        data.append(['Tardanzas:', str(late), ''])
        data.append(['Tasa de asistencia:', f'{attendance_rate: .1f}%', ''])

        table = Table(data, colWidths=[2 * inch, 2 * inch, 3 * inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF0000')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, total), 1, colors.black),
            ('BACKGROUND', (0, 1), (-1, total), colors.beige),
        ]))

        elements.append(table)
        doc.build(elements)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/pdf')
        filename = f'asistencia_{student.username}_{datetime.now().strftime("%Y%m%d")}.pdf'
        response['Content-Disposition'] = 'attachment' + '\x3B' + f' filename="{filename}"'

        return response


class ExcelReportGenerator:
    """
    Generador de reportes en Excel usando pandas y openpyxl.
    """

    @staticmethod
    def generate_grades_excel(grades, filename='calificaciones'):
        """
        Generar reporte de calificaciones en Excel.

        Args:
            grades: QuerySet de calificaciones
            filename: Nombre del archivo sin extensión

        Returns:
            HttpResponse con el archivo Excel
        """
        # Preparar datos
        data = []
        for grade in grades:
            data.append({
                'Estudiante': grade.student.get_full_name(),
                'Código': grade.student.username,
                'Materia': grade.subject.name,
                'Curso': grade.subject.course.name,
                'Tipo': grade.get_grade_type_display(),
                'Calificación': float(grade.value),
                'Peso (%)': float(grade.weight),
                'Calificado por': grade.graded_by.get_full_name() if grade.graded_by else '-',
                'Fecha': grade.graded_date.strftime('%Y-%m-%d'),
                'Estado': 'Aprobado' if grade.is_passing else 'Reprobado',
                'Letra': grade.letter_grade
            })

        # Crear DataFrame
        df = pd.DataFrame(data)

        # Crear archivo Excel en memoria
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Calificaciones')

            # Obtener el workbook y worksheet para dar formato
            writer.book
            worksheet = writer.sheets['Calificaciones']

            # Formato del encabezado
            header_fill = PatternFill(
                start_color='FF0000',
                end_color='FF0000',
                fill_type='solid',
            )
            header_font = Font(color='FFFFFF', bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except BaseException:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)

        # Crear respuesta HTTP
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        date_str = datetime.now().strftime("%Y%m%d")
        response['Content-Disposition'] = 'attachment' + '\x3B' + f' filename="{filename}_{date_str}.xlsx"'

        return response

    @staticmethod
    def generate_attendance_excel(attendances, filename='asistencias'):
        """
        Generar reporte de asistencias en Excel.
        """
        data = []
        for att in attendances:
            data.append({
                'Estudiante': att.student.get_full_name(),
                'Código': att.student.username,
                'Curso': att.course.name,
                'Fecha': att.date.strftime('%Y-%m-%d'),
                'Estado': att.get_status_display(),
                'Notas': att.notes or '-',
                'Registrado por': att.recorded_by.get_full_name() if att.recorded_by else '-'
            })

        df = pd.DataFrame(data)
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Asistencias')

            writer.book
            worksheet = writer.sheets['Asistencias']

            # Formato
            header_fill = PatternFill(
                start_color='FF0000',
                end_color='FF0000',
                fill_type='solid',
            )
            header_font = Font(color='FFFFFF', bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except BaseException:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        date_str = datetime.now().strftime("%Y%m%d")
        response['Content-Disposition'] = 'attachment' + '\x3B' + f' filename="{filename}_{date_str}.xlsx"'

        return response
